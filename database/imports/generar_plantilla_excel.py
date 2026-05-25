#!/usr/bin/env python3
"""Genera plantilla Excel con listas desplegables. Ejecutar: py -3 database/imports/generar_plantilla_excel.py"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).parent / "restaurantes_lambayeque_plantilla.xlsx"

HEADERS = [
    "nombre",
    "especialidad_gastronomica",
    "categoria_establecimiento",
    "entorno_restaurante",
    "ambiente_restaurante",
    "rango_precios",
    "momento_recomendado",
    "servicios",
    "ubicacion",
    "publico_objetivo",
    "dias",
    "hora_apertura",
    "hora_cierre",
    "direccion",
    "latitud",
    "longitud",
    "descripcion_corta",
    "telefono",
]

CATALOGOS = {
    "especialidad_gastronomica": sorted(
        {
            "Criolla",
            "Criolla peruana",
            "Marina",
            "Ceviche",
            "Chifa",
            "Lambayecana",
            "Tradicional lambayecana",
            "Internacional",
            "Fusión",
            "Pollos y parrillas",
            "Postres y cafetería",
            "Vegetariana",
            "Coctelería",
            "Comida regional",
            "Comida rápida",
            "Parrilla",
        }
    ),
    "categoria_establecimiento": sorted(
        {
            "Restaurante tradicional",
            "Gourmet",
            "Pollería",
            "Cevichería",
            "Restobar",
            "Comida rápida",
            "Cafetería",
            "Parrilla",
            "Pizzería",
            "Chifa",
            "Huarique",
            "Familiar",
        }
    ),
    "entorno_restaurante": [
        "Urbano",
        "Campestre",
        "Centro histórico",
        "Vista al mar",
        "Centro comercial",
    ],
    "ambiente_restaurante": sorted(
        {
            "Casual",
            "Elegante",
            "Moderno",
            "Tradicional",
            "Familiar",
            "Romántico",
            "Cultural",
            "Festivo",
            "Tranquilo",
        }
    ),
    "rango_precios": ["economico", "moderado", "premium"],
    "momento_recomendado": ["Desayuno", "Almuerzo", "Cena", "Brunch", "Bar"],
    "servicios": sorted(
        {
            "WiFi",
            "Estacionamiento",
            "Delivery",
            "Reservas",
            "Terraza",
            "Música en vivo",
            "Acceso silla de ruedas",
            "Acepta mascotas",
            "Pago con tarjeta",
            "Para llevar",
        }
    ),
    "ubicacion": ["Chiclayo", "Lambayeque", "Pimentel", "La Victoria", "Eten", "Monsefú", "Ferreñafe"],
    "publico_objetivo": sorted(
        {
            "Turistas",
            "Familias",
            "Familia",
            "Grupos",
            "Parejas",
            "Pareja",
            "Jóvenes",
            "Ejecutivos",
            "Amigos",
            "Negocios",
            "Solo",
        }
    ),
    "dias": ["lun-dom", "lun-vie", "lun,mie,vie", "sab,dom"],
}

EJEMPLOS = [
    [
        "Urumanka Barra Restaurante",
        "Criolla",
        "Restaurante tradicional",
        "Urbano",
        "Elegante",
        "moderado",
        "Almuerzo",
        "WiFi|Reservas",
        "Chiclayo",
        "Turistas",
        "lun-dom",
        "07:00",
        "21:00",
        "Av. Ejemplo 123, Chiclayo",
        -6.77137,
        -79.84088,
        "Sabores criollos en el corazón de Chiclayo",
        "+51999999999",
    ],
    [
        "La Casona Criolla",
        "Criolla",
        "Restaurante tradicional",
        "Centro histórico",
        "Casual",
        "moderado",
        "Almuerzo|Cena",
        "WiFi|Delivery",
        "Chiclayo",
        "Familia|Turistas",
        "lun-dom",
        "11:00",
        "22:00",
        "Calle Izaga 250",
        "",
        "",
        "",
        "",
    ],
]

# Columnas con lista desplegable (índice 1-based en Excel)
DROPDOWN_COLS = {
    2: "especialidad_gastronomica",
    3: "categoria_establecimiento",
    4: "entorno_restaurante",
    5: "ambiente_restaurante",
    6: "rango_precios",
    7: "momento_recomendado",
    9: "ubicacion",
    11: "dias",
}


def main() -> None:
    wb = Workbook()
    ws_cat = wb.active
    ws_cat.title = "Catalogos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C41E3A")

    for col_idx, (key, values) in enumerate(CATALOGOS.items(), start=1):
        cell = ws_cat.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        for row_idx, val in enumerate(values, start=2):
            ws_cat.cell(row=row_idx, column=col_idx, value=val)
        letter = get_column_letter(col_idx)
        ws_cat.column_dimensions[letter].width = max(18, len(key) + 2)

    ws = wb.create_sheet("Restaurantes", 0)
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 16 if col_idx > 1 else 28

    ws.freeze_panes = "A2"

    row = 2
    for ejemplo in EJEMPLOS:
        for col_idx, val in enumerate(ejemplo, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        row += 1

    max_rows = 120
    last_data_row = max_rows

    for col_idx, cat_key in DROPDOWN_COLS.items():
        values = CATALOGOS[cat_key]
        cat_col = list(CATALOGOS.keys()).index(cat_key) + 1
        cat_letter = get_column_letter(cat_col)
        last_cat_row = len(values) + 1
        formula = f"=Catalogos!${cat_letter}$2:${cat_letter}${last_cat_row}"
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error = "Elige un valor de la lista en la hoja Catalogos"
        dv.errorTitle = "Valor no válido"
        dv.prompt = "Selecciona de la lista"
        dv.promptTitle = cat_key
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_data_row}")

    ws_inst = wb.create_sheet("Instrucciones")
    instrucciones = [
        "PLANTILLA — Restaurantes Lambayeque (DiscoverLambo)",
        "",
        "1. Rellena solo la hoja «Restaurantes» (filas 2 en adelante).",
        "2. Usa las listas desplegables donde aparecen (columnas con catálogo).",
        "3. Varios valores en una celda: separa con |  →  WiFi|Reservas|Delivery",
        "4. momento_recomendado y publico_objetivo: también con | si son varios",
        "5. rango_precios: solo economico, moderado o premium (sin tildes)",
        "6. dias: lun-dom = lunes a domingo",
        "7. latitud y longitud: necesarios para mapa y rutas IA",
        "8. Puedes copiar/pegar desde Google Sheets; conserva las columnas.",
        "9. Guarda este archivo y luego impórtalo con: php artisan restaurants:import",
        "",
        "Hoja «Catalogos»: no borres; alimenta las listas desplegables.",
    ]
    for i, line in enumerate(instrucciones, start=1):
        ws_inst.cell(row=i, column=1, value=line)
    ws_inst.column_dimensions["A"].width = 80

    wb.save(OUTPUT)
    print(f"Generado: {OUTPUT}")


if __name__ == "__main__":
    main()
